###Dictionaries/lists###
import openpyxl
from openpyxl import load_workbook


class hAppet:

    prisLista = {}

    staticmethod
    def dataUt():
        wb = load_workbook('prislista.xlsx', read_only=True)
        sheet = wb.active
        mxrw = sheet.max_row #alla rows med innehåll

        adict = {(sheet.cell(row=i, column=1).value): (sheet.cell(row=i, column=2).value) for i in range(2,mxrw)}

        return adict

    antalLista = {
        "Sambusa": 0,"Ostrullar": 0,"Hummus": 0,
    }

    saldLista = {
        "Sambusa": 0,"Ostrullar": 0,"Hummus": 0,

    }

    inkoLista = {
        "Sambusa": 0,"Ostrullar": 0,"Hummus": 0,

    }



    masterlist = [antalLista, saldLista, inkoLista]



    ###Funktioner###
    staticmethod
    def svdt(x):

        wb = openpyxl.Workbook()
        sheet = wb.active
        mxrw = sheet.max_row

        for i, entry in enumerate(x, start=1): sheet.cell(row=2, column=i, value = entry)

        wb.save('usrdata.xlsx')


    staticmethod
    def inko():
        for l in masterlist:
            for ratt in l:
                
                print ('{} {} salda for {} kr'.format(saldLista[ratt], ratt, inkoLista[ratt]))

            break


    staticmethod
    def bestall(best):
        betSum = 0
        for ratt in best:


            hAppet.antalLista[ratt] = int(input('\nAntal {} '.format(ratt)))
            betSum += hAppet.prisLista[ratt] * hAppet.antalLista[ratt]    
            
            hAppet.saldLista[ratt] += hAppet.antalLista[ratt]
            hAppet.inkoLista[ratt] += hAppet.saldLista[ratt] * hAppet.prisLista[ratt]
            
            
        print ('\nSumma bestallning..........:' , betSum, 'kr\n')



    staticmethod
    def prisMeny1():
        print ("\nPrislista\n")

        prisLista = hAppet.dataUt()

        for k,v in prisLista.items():

            print (k + ': {}kr'.format(v))


        hAppet.bestall(hAppet.antalLista)



    ###Valkommen###



while True:
    

    print ('Welcome to Hoppet!\nMeny[1]\nIrak[2]\nMatlagnings kurs[3]\nForsaljning\nAvsluta\n\nKontakt: 072xxxxxxx\n')
    val = input('Vad kan vi hjalpa med?\n ')


    for v in val:
        
        if '1' in v:
            hAppet.prisMeny1()        
        elif '2' in v:
            print ('500kr/person\n')
        elif '3' in v:
            print ('750kr/person\n')
        elif '4' in v:
            inko()
            print ('\n')
        elif '5' in v:
            print ('Tack och hej!')
            exit()
        
